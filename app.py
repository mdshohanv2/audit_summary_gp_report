import streamlit as st
import pandas as pd
from datetime import datetime

# Page configuration
st.set_page_config(
    page_title="Audit Analyzer",
    page_icon="📊",
    layout="wide"
)

# Minimal Shadcn-like CSS
st.markdown("""
    <style>
    /* Main container styling */
    .block-container {
        padding-top: 3rem;
        padding-bottom: 3rem;
    }
    
    /* Typography */
    h1 {
        font-weight: 700;
        letter-spacing: -0.025em;
    }
    
    /* Subheader spacing */
    .stSubheader {
        margin-top: 1.5rem;
        font-weight: 600;
    }

    /* Minimalist upload styling with theme-aware borders */
    .stFileUploader {
        border: 1px solid rgba(128, 128, 128, 0.2);
        border-radius: 8px;
        padding: 10px;
    }

    /* Footer/Info styling */
    .stAlert {
        border-radius: 8px;
        border: 1px solid rgba(128, 128, 128, 0.2);
        padding: 0.75rem 1rem;
        min-height: 52px;
        display: flex;
        align-items: center;
    }
    
    /* Match Download button style to Alert box */
    .stDownloadButton, .stDownloadButton > button {
        width: 100%;
        min-height: 52px;
        border-radius: 8px !important;
        border: 1px solid rgba(128, 128, 128, 0.2) !important;
        background-color: transparent !important;
        transition: all 0.2s ease;
    }
    .stDownloadButton > button:hover {
        background-color: rgba(128, 128, 128, 0.1) !important;
        border-color: rgba(128, 128, 128, 0.4) !important;
    }
    
    /* Divider styling */
    hr {
        margin-top: 2rem;
        margin-bottom: 2rem;
        opacity: 0.1;
    }
    </style>
    """, unsafe_allow_html=True)

from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import io

# Database Parameters from ME_MTD_Report_Automation.py
DB_PARAMS = {
    "dbname": "gp_dev",
    "user": "report_user",
    "password": "report#Gp*User!#__D",
    "host": "gp-stg.cf44ysgum7u8.ap-southeast-1.rds.amazonaws.com",
    "port": 5432
}
DATABASE_URL = f"postgresql://{DB_PARAMS['user']}:{DB_PARAMS['password']}@{DB_PARAMS['host']}:{DB_PARAMS['port']}/{DB_PARAMS['dbname']}"
engine = create_engine(DATABASE_URL)

def get_me_mtd_query(from_date, custom_date):
    return text(f"""
    WITH report_data AS (
        SELECT from_date, to_date, circle, region, cluster_name, territory, house, dh_code, me_code,
        CASE WHEN successful_visit > 0 THEN 'Working' ELSE 'Absent' END AS status,
        mtd_working_day, COALESCE(NULLIF(monthly_target, 0), 0) AS monthly_target,
        COALESCE(NULLIF(open, 0), 0) AS open, COALESCE(NULLIF(temporarily_closed, 0), 0) AS temporarily_closed,
        COALESCE(NULLIF(permanently_closed, 0), 0) AS permanently_closed, COALESCE(NULLIF(moved, 0), 0) AS moved,
        COALESCE(NULLIF(not_found, 0), 0) AS not_found, avg_time_spent_in_between_pos_service,
        COALESCE(NULLIF(mtd_target, 0), 0) AS mtd_target, COALESCE(NULLIF(mtd_served_days, 0), 0) AS mtd_served_days,
        COALESCE(NULLIF(late_checkin, 0), 0) AS late_checkin, COALESCE(NULLIF(successful_visit, 0), 0) AS successful_visit,
        COALESCE(ROUND((successful_visit::numeric / NULLIF(mtd_target, 0)) * 100, 0), 0)::int || '%' AS successful_visit_ach,
        avg_time_spent_in_pos, COALESCE(NULLIF(three_or_more_min_visit_count, 0), 0) AS three_or_more_min_visit_count,
        '80%' AS grater_or_equ_three_min_visit_target_percent,
        COALESCE(ROUND((three_or_more_min_visit_count ::numeric / NULLIF(mtd_target, 0)) * 100, 0), 0)::int || '%' AS less_equ_three_min_visit_ach,
        COALESCE(NULLIF(monthly_unique_pos_visit_target, 0), 0) AS monthly_unique_pos_visit_target,
        COALESCE(NULLIF(one_successful_visit_count_pos, 0), 0) AS one_successful_visit_count_pos,
        COALESCE(NULLIF(two_successful_visit_count_pos, 0), 0) AS two_successful_visit_count_pos,
        COALESCE(NULLIF(three_or_more_successful_visit_count_pos, 0),0) AS three_or_more_successful_visit_count_pos,
        COALESCE(NULLIF(COALESCE(one_successful_visit_count_pos, 0) + COALESCE(two_successful_visit_count_pos, 0) + COALESCE(three_or_more_successful_visit_count_pos, 0), 0), 0) AS mtd_unique_pos_visit,
        COALESCE(NULLIF(COALESCE(five_or_more_poster_pos, 0) + COALESCE(four_poster_pos, 0) + COALESCE(three_poster_pos, 0), 0), 0) AS three_or_more_poster_pos_count,
        '90%' AS three_or_more_poster_pos_target, COALESCE(NULLIF(five_or_more_poster_pos, 0), 0) AS five_or_more_poster_pos,
        COALESCE(NULLIF(four_poster_pos, 0), 0) AS four_poster_pos, COALESCE(NULLIF(three_poster_pos, 0), 0) AS three_poster_pos,
        COALESCE(NULLIF(two_poster_pos, 0), 0) AS two_poster_pos, COALESCE(NULLIF(one_poster_pos, 0), 0) AS one_poster_pos,
        COALESCE(NULLIF(COALESCE(five_or_more_poster_pos, 0) + COALESCE(four_poster_pos, 0) + COALESCE(three_poster_pos, 0) + COALESCE(two_poster_pos, 0) + COALESCE(one_poster_pos, 0), 0), 0) AS all_poster_pos,
        COALESCE(NULLIF(COALESCE(three_or_more_festoon_pos, 0) + COALESCE(two_festoon_pos, 0) + COALESCE(one_festoon_pos, 0), 0), 0) AS atleast_one_festoon_pos_count,
        '70%' AS festoon_pos_count_target, COALESCE(NULLIF(three_or_more_festoon_pos, 0), 0) AS three_or_more_festoon_pos,
        COALESCE(NULLIF(two_festoon_pos, 0), 0) AS two_festoon_pos, COALESCE(NULLIF(one_festoon_pos, 0), 0) AS one_festoon_pos,
        COALESCE(NULLIF(shopscreen_pos, 0), 0) AS shopscreen_pos, '30%' AS shopscreen_pos_count_target,
        COALESCE(NULLIF(cover_sticker_pos, 0), 0) AS cover_sticker_pos, '30%' AS cover_sticker_pos_count_target,
        COALESCE(NULLIF(no_bag, 0), 0) AS no_bag, COALESCE(NULLIF(damaged_bag, 0), 0) AS damaged_bag,
        COALESCE(NULLIF(no_hammer, 0), 0) AS no_hammer, COALESCE(NULLIF(no_pliers, 0), 0) AS no_pliers,
        COALESCE(NULLIF(anti_cutter, 0), 0) AS anti_cutter, COALESCE(NULLIF(duster, 0), 0) AS duster,
        COALESCE(NULLIF(scotch_tape, 0), 0) AS scotch_tape, COALESCE(NULLIF(glue, 0), 0) AS glue,
        COALESCE(NULLIF(tar_or_string, 0), 0) AS tar_or_string, COALESCE(NULLIF(board_pin, 0), 0) AS board_pin,
        COALESCE(NULLIF(pin_or_perek, 0), 0) AS pin_or_perek
        FROM dashboard_reporting.me_mtd_data
        WHERE to_date = '{custom_date}' AND from_date = '{from_date}'
    )
    SELECT from_date AS "From Date", to_date AS "To Date", circle AS "Circle", region AS "Region",
    cluster_name AS "Cluster", territory AS "Territory", house AS "Distribution House", dh_code AS "DH Code",
    me_code AS "ME Code", status as "Status", mtd_working_day as "MTD Working Days Count", mtd_served_days AS "MTD Served Days",
    CASE WHEN COALESCE(mtd_working_day, 0) - COALESCE(mtd_served_days, 0) < 0 THEN 0
    ELSE COALESCE(mtd_working_day, 0) - COALESCE(mtd_served_days, 0) END AS "Vacant Days",
    late_checkin AS "Late Check In Days Count [After 9 AM]", monthly_target AS "Monthly Visit Target",
    mtd_target AS "MTD Visit Target", successful_visit AS "MTD Successful Visits",
    successful_visit_ach as "Successful Visit Ach%", open AS "Visit Status Open",
    temporarily_closed AS "Visit Status Temporary Closed", permanently_closed AS "Visit Status Permanently Closed",
    moved AS "Visit Status Moved", not_found AS "Visit Status Not Found",
    avg_time_spent_in_pos AS "Avg Spent Time in POS", avg_time_spent_in_between_pos_service AS "Avg Time Spent in Between POS Service",
    three_or_more_min_visit_count AS ">= 3 Min Visit Count", grater_or_equ_three_min_visit_target_percent as ">= 3 Min Visit Target %",
    less_equ_three_min_visit_ach as ">= 3 Min Visit Ach%", monthly_unique_pos_visit_target as "Monthly Unique POS Visit Target",
    one_successful_visit_count_pos as "1 Successful Visit Count POS", two_successful_visit_count_pos as "2 Successful Visit Count POS",
    three_or_more_successful_visit_count_pos as "3&+ Successful Visit Count POS", mtd_unique_pos_visit as "MTD Unique POS Visited",
    COALESCE(ROUND((mtd_unique_pos_visit ::numeric / NULLIF(monthly_unique_pos_visit_target, 0)) * 100, 0), 0)::int || '%' AS "MTD Unique POS Visit Ach.%",
    three_or_more_poster_pos_count as ">= 3 Poster POS Count", three_or_more_poster_pos_target as ">= 3 Poster POS Count Target",
    COALESCE(ROUND((three_or_more_poster_pos_count ::numeric / NULLIF(monthly_unique_pos_visit_target, 0)) * 100, 0), 0)::int || '%' as ">= 3 Poster POS Count Ach%",
    five_or_more_poster_pos as "5&+ Poster POS", four_poster_pos as "4 Poster POS", three_poster_pos as "3 Poster POS",
    two_poster_pos as "2 Poster POS", one_poster_pos as "1 Poster POS",
    COALESCE(NULLIF(mtd_unique_pos_visit - all_poster_pos, 0), 0) AS "0 Poster POS",
    atleast_one_festoon_pos_count as "Atleast 1 Festoon POS Count", festoon_pos_count_target as "Festoon POS Count Target",
    COALESCE(ROUND((atleast_one_festoon_pos_count ::numeric / NULLIF(monthly_unique_pos_visit_target, 0)) * 100, 0), 0)::int || '%' "Festoon POS Count Ach%",
    three_or_more_festoon_pos as "3&+ Festoon POS", two_festoon_pos as "2 Festoon POS", one_festoon_pos as "1 Festoon POS",
    COALESCE(NULLIF(mtd_unique_pos_visit - atleast_one_festoon_pos_count, 0), 0) as "0 Festoon POS",
    shopscreen_pos as "Atleast 1 Shopscreen POS Count", shopscreen_pos_count_target as "Shopscreen POS Count Target",
    COALESCE(ROUND((shopscreen_pos ::numeric / NULLIF(monthly_unique_pos_visit_target, 0)) * 100, 0), 0)::int || '%' as "Shopscreen POS Count Ach%",
    cover_sticker_pos as "Atleast 1 Cover Sticker POS Count", cover_sticker_pos_count_target as "Cover Sticker POS Count Target",
    COALESCE(ROUND((cover_sticker_pos ::numeric / NULLIF(monthly_unique_pos_visit_target, 0)) * 100, 0), 0)::int || '%' as "Cover Sticker POS Count Ach%",
    no_bag AS "No Bag", damaged_bag AS "Damaged Bag", no_hammer AS "No Hammer", no_pliers AS "No Pliers",
    anti_cutter AS "Anti Cutter", duster AS "Duster", scotch_tape AS "Scotch Tape", glue AS "Glue",
    tar_or_string AS "Tar/String", board_pin AS "Board Pin", pin_or_perek AS "Pin/Perek"
    FROM report_data order by me_code
    """)

def get_sup_mtd_query(custom_date):
    return text(f"""
    WITH report_data AS (
        SELECT sm.from_date, sm.to_date, sm.circle, sm.region, sm.sup_code,
        COALESCE(sm.mtd_working_day, 0) AS mtd_working_day, COALESCE(sm.sup_served_days, 0) AS sup_served_days,
        GREATEST(COALESCE(sm.mtd_working_day, 0) - COALESCE(sm.sup_served_days, 0), 0) AS vacant_days,
        COALESCE(sm.late_check_in_count, 0) AS late_check_in_count, sm.late_check_in_average_time,
        COALESCE(sm.mtd_me_tag_target, 0) AS mtd_me_tag_target, COALESCE(sm.me_tag_ach, 0) AS me_tag_ach,
        COALESCE(ROUND((sm.me_tag_ach::numeric / NULLIF(sm.mtd_me_tag_target, 0)) * 100, 0), 0)::int || '%' AS joint_ach,
        COALESCE(sm.mtd_working_day, 0) * 20 AS mtd_joint_visit_target, COALESCE(sm.mtd_joint_visit_ach, 0) AS mtd_joint_visit_ach,
        COALESCE(sm.successful_joint_visit_ach, 0) AS successful_joint_visit_ach,
        COALESCE(sm.mtd_working_day, 0) * 10 AS mtd_solo_visit_terget, COALESCE(sm.mtd_solo_visit_ach, 0) AS mtd_solo_visit_ach,
        COALESCE(sm.visit_status_open, 0) AS visit_status_open, COALESCE(sm.visit_status_same_pos_different_name, 0) AS visit_status_same_pos_different_name,
        COALESCE(sm.mtd_unique_pos_visit, 0) AS mtd_unique_pos_visit,
        COALESCE(sm.mtd_joint_visit_ach, 0) + COALESCE(sm.mtd_solo_visit_ach, 0) AS mtd_total_pos_visit,
        COALESCE(sm.mtd_me_successful_visit, 0) AS mtd_me_successful_visit, COALESCE(sm.mtd_me_unique_pos_visited, 0) AS mtd_me_unique_pos_visited,
        COALESCE(sm.three_min_iseqgreater_visit_count, 0) AS three_min_iseqgreater_visit_count,
        '80%' AS three_min_visit_target, COALESCE(sm.three_iseqgreater_poster_pos_count, 0) AS three_iseqgreater_poster_pos_count,
        '90%' AS poster_pos_count_target, COALESCE(sm.five_plus_poster_pos, 0) AS five_plus_poster_pos,
        COALESCE(sm.four_poster_pos, 0) AS four_poster_pos, COALESCE(sm.three_poster_pos, 0) AS three_poster_pos,
        COALESCE(sm.two_poster_pos, 0) AS two_poster_pos, COALESCE(sm.one_poster_pos, 0) AS one_poster_pos,
        COALESCE(sm.five_plus_poster_pos, 0) + COALESCE(sm.four_poster_pos, 0) + COALESCE(sm.three_poster_pos, 0) + 
        COALESCE(sm.two_poster_pos, 0) + COALESCE(sm.one_poster_pos, 0) AS all_poster_pos,
        '70%' AS festoon_pos_count_target, COALESCE(sm.three_plus_festoon_pos, 0) AS three_plus_festoon_pos,
        COALESCE(sm.two_festoon_pos, 0) AS two_festoon_pos, COALESCE(sm.one_festoon_pos, 0) AS one_festoon_pos,
        COALESCE(sm.three_plus_festoon_pos, 0) + COALESCE(sm.two_festoon_pos, 0) + COALESCE(sm.one_festoon_pos, 0) AS all_festoon_pos,
        COALESCE(sm.atleast_one_shopscreen_pos, 0) AS atleast_one_shopscreen_pos, '30%' AS shopscreen_pos_count_target,
        COALESCE(sm.atleast_one_cover_sticker_pos, 0) AS atleast_one_cover_sticker_pos, '30%' AS cover_sticker_pos_count_target
        FROM dashboard_reporting.sup_mtd_data sm
        WHERE sm.to_date = '{custom_date}'
    )
    SELECT from_date AS "Start Date", to_date AS "Till Date", circle AS "Circle", region AS "Region", sup_code AS "Sup Code",
    mtd_working_day AS "MTD Working Day Count", sup_served_days AS "SUP Served Days", vacant_days AS "Vacant Days",
    late_check_in_count AS "Late Check In Count", late_check_in_average_time AS "Late Check in Average Time",
    mtd_me_tag_target AS "Monthly ME Tag Target", me_tag_ach AS "ME Tag Ach", joint_ach AS "Ach %",
    mtd_joint_visit_target AS "MTD Joint Visit Target", mtd_joint_visit_ach AS "MTD Joint Visit Ach",
    successful_joint_visit_ach AS "MTD Successful Joint Visit Ach",
    COALESCE(ROUND((mtd_joint_visit_ach::numeric / NULLIF(mtd_joint_visit_target, 0)) * 100, 0), 0)::int || '%' AS "Final Ach %",
    mtd_solo_visit_terget AS "MTD Solo Visit Target", mtd_solo_visit_ach AS "MTD Solo Visit Ach",
    COALESCE(ROUND((mtd_solo_visit_ach::numeric / NULLIF(mtd_solo_visit_terget, 0)) * 100, 0), 0)::int || '%' AS "Ach %",
    visit_status_open AS "Visit Status Open", visit_status_same_pos_different_name AS "Visit Status Same POS Different Name",
    mtd_total_pos_visit AS "MTD Total POS Visit", mtd_unique_pos_visit AS "MTD Unique POS Visit",
    COALESCE(ROUND((mtd_unique_pos_visit::numeric / NULLIF(mtd_total_pos_visit, 0)) * 100, 0), 0)::int || '%' AS "MTD Unique POS Visit %",
    three_min_iseqgreater_visit_count AS ">= 3 Min Visit Count", three_min_visit_target AS ">= 3 Min Visit Target %",
    COALESCE(ROUND((three_min_iseqgreater_visit_count::numeric / NULLIF(mtd_me_successful_visit, 0)) * 100, 0), 0)::int || '%' AS ">= 3 Min Visit Ach%",
    three_iseqgreater_poster_pos_count AS ">= 3 Poster POS Count", poster_pos_count_target AS ">= 3 Poster POS Count Target",
    COALESCE(ROUND((three_iseqgreater_poster_pos_count::numeric / NULLIF(mtd_me_unique_pos_visited, 0)) * 100, 0), 0)::int || '%' AS ">= 3 Poster POS Count Ach%",
    five_plus_poster_pos AS "5&+ Poster POS", four_poster_pos AS "4 Poster POS", three_poster_pos AS "3 Poster POS",
    two_poster_pos AS "2 Poster POS", one_poster_pos AS "1 Poster POS",
    COALESCE(mtd_me_unique_pos_visited, 0) - COALESCE(all_poster_pos, 0) AS "0 Poster POS",
    all_festoon_pos AS "Atleast 1 Festoon POS Count", festoon_pos_count_target AS "Festoon POS Count Target",
    COALESCE(ROUND((all_festoon_pos::numeric / NULLIF(mtd_me_unique_pos_visited, 0)) * 100, 0), 0)::int || '%' AS "Festoon POS Count Ach%",
    three_plus_festoon_pos AS "3&+ Festoon POS", two_festoon_pos AS "2 Festoon POS", one_festoon_pos AS "1 Festoon POS",
    COALESCE(mtd_me_unique_pos_visited, 0) - COALESCE(all_festoon_pos, 0) AS "0 Festoon POS",
    atleast_one_shopscreen_pos AS "Atleast 1 Shopscreen POS Count", shopscreen_pos_count_target AS "Shopscreen POS Count Target",
    COALESCE(ROUND((atleast_one_shopscreen_pos::numeric / NULLIF(mtd_me_unique_pos_visited, 0)) * 100, 0), 0)::int || '%' AS "Shopscreen POS Count Ach%",
    atleast_one_cover_sticker_pos AS "Atleast 1 Cover Sticker POS Count", cover_sticker_pos_count_target AS "Cover Sticker POS Count Target",
    COALESCE(ROUND((atleast_one_cover_sticker_pos::numeric / NULLIF(mtd_me_unique_pos_visited, 0)) * 100, 0), 0)::int || '%' AS "Cover Sticker POS Count Ach%"
    FROM report_data;
    """)

def fetch_db_data(query):
    try:
        df = pd.read_sql_query(query, con=engine)
        return df
    except Exception as e:
        st.error(f"Error fetching data: {e}")
        return pd.DataFrame()

def process_me_summary(df):
    if df.empty: return pd.DataFrame()
    circle_col, region_col, assigned_me_col, status_col = 'Circle', 'Region', 'ME Code', 'Status'
    summary_df = df[[circle_col, region_col]].drop_duplicates().sort_values(by=circle_col).reset_index(drop=True)
    summary_df['Assigned ME'] = summary_df[region_col].apply(lambda r: df[df[region_col] == r][assigned_me_col].nunique())
    summary_df['Working ME'] = summary_df[region_col].apply(lambda r: df[(df[region_col] == r) & (df[status_col] == 'Working')][assigned_me_col].nunique())
    status_map = {'Visit Status Open': 'Open', 'Visit Status Temporary Closed': 'Temporary closed', 'Visit Status Permanently Closed': 'Permanently closed', 'Visit Status Moved': 'Moved', 'Visit Status Not Found': 'Not found'}
    for raw, display in status_map.items(): summary_df[display] = summary_df[region_col].apply(lambda r: df[df[region_col] == r][raw].sum())
    summary_df['POS Visited'] = summary_df[list(status_map.values())].sum(axis=1)
    summary_df['Open POS %'] = summary_df['Open'] / summary_df['POS Visited']
    summary_df['Unique POS Visited'] = summary_df[region_col].apply(lambda r: df[df[region_col] == r]['MTD Unique POS Visited'].sum())
    poster_map = {'Min 1 Festoon': 'Atleast 1 Festoon POS Count', 'Min 1 Shopscreen': 'Atleast 1 Shopscreen POS Count', 'Cover Sticker': 'Atleast 1 Cover Sticker POS Count', '0 Poster': '0 Poster POS', '1 Poster': '1 Poster POS', '2 Poster': '2 Poster POS', '3 and + Poster': '>= 3 Poster POS Count'}
    for display, raw in poster_map.items(): summary_df[display] = summary_df[region_col].apply(lambda r: df[df[region_col] == r][raw].sum() if raw in df.columns else 0)
    for col in poster_map.keys(): summary_df[col + ' POS %'] = summary_df[col] / summary_df['Unique POS Visited']
    cols = ['Circle', 'Region', 'Assigned ME', 'Working ME', 'POS Visited', 'Open', 'Temporary closed', 'Permanently closed', 'Moved', 'Not found', 'Open POS %', 'Unique POS Visited', 'Min 1 Festoon', 'Min 1 Shopscreen', 'Cover Sticker', '0 Poster', '1 Poster', '2 Poster', '3 and + Poster', 'Min 1 Festoon POS %', 'Min 1 Shopscreen POS %', 'Cover Sticker POS %', '0 Poster POS %', '1 Poster POS %', '2 Poster POS %', '3 and + Poster POS %']
    summary_df = summary_df[cols].copy()
    num_cols = [c for c in cols if c not in ['Circle', 'Region'] and not c.endswith('%')]
    pct_cols = [c for c in summary_df.columns if c.endswith('%')]
    total_row = {c: summary_df[c].sum() if c in num_cols else '' for c in cols}
    for c in pct_cols:
        total_row[c] = f"{round(summary_df[c].mean() * 100):.0f} %"
        summary_df[c] = (summary_df[c] * 100).round(0).astype(int).astype(str) + " %"
    return pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)

def generate_automation_excel(df_me, df_summary, df_sup, date_filename):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_me.to_excel(writer, sheet_name="ME MTD RPT", index=False)
        df_summary.to_excel(writer, sheet_name="National Summary", index=False)
        if not df_sup.empty: df_sup.to_excel(writer, sheet_name="SUP MTD RPT", index=False)
        
        from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
        from openpyxl.formatting.rule import ColorScaleRule
        thin = Side(border_style="thin", color="000000")
        all_border = Border(top=thin, bottom=thin, left=thin, right=thin)
        
        for sheet in writer.book.worksheets:
            sheet.row_dimensions[1].height = 60
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row: cell.border = all_border
            for cell in sheet[1]:
                cell.font = Font(bold=True); cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                sheet.column_dimensions[cell.column_letter].width = 15

            # Convert % text to actual percentages for proper Excel behavior
            if sheet.title in ["ME MTD RPT", "National Summary"]:
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    for cell in row:
                        if isinstance(cell.value, str) and '%' in cell.value:
                            try:
                                cell.value = float(cell.value.replace('%', '')) / 100
                                cell.number_format = '0%'
                            except: pass

            # Conditional Formatting for National Summary
            if sheet.title == "National Summary":
                last_row = sheet.max_row
                for col_letter in ['K', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
                    col_range = f"{col_letter}2:{col_letter}{last_row-1}"
                    try:
                        sheet.conditional_formatting.add(col_range, ColorScaleRule(
                            start_type='min', start_color='F8696B',
                            mid_type='percentile', mid_value=50, mid_color='FFEB84',
                            end_type='max', end_color='63BE7B'
                        ))
                    except: pass
    return output.getvalue()

# Helper function to process Shinsa report
def get_shinsa_summary(file):
    try:
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            # For Excel files, we might need openpyxl
            df = pd.read_excel(file)
        
        # Look for the target column (case-insensitive)
        target_col = 'visit_pos_status'
        actual_col = next((c for c in df.columns if str(c).strip().lower() == target_col), None)
        
        if actual_col:
            # Normalize data: lowercase, stripped, handle nulls
            status_series = df[actual_col].astype(str).str.strip().str.lower()
            
            # Define what counts as "empty"
            empty_markers = ['nan', '', 'none', 'null']
            status_series = status_series.apply(lambda x: 'unknown/empty' if x in empty_markers else x)
            
            # Count occurrences
            counts = status_series.value_counts().reset_index()
            counts.columns = ['Status', 'Count']
            
            # Calculate total
            total = counts['Count'].sum()
            
            return counts, total, None
        else:
            return None, 0, f"Column '{target_col}' not found in the uploaded file."
    except Exception as e:
        return None, 0, f"Error processing file: {str(e)}"

# Helper function to process MTD report
def get_mtd_summary(file):
    try:
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        # 1. Total Successful Visits
        target_col = 'mtd successful visits'
        actual_col = next((c for c in df.columns if str(c).strip().lower().replace('\n', ' ') == target_col), None)
        total_sum = pd.to_numeric(df[actual_col], errors='coerce').sum() if actual_col else 0

        # 2. Individual Status Sums (New Requirement)
        status_mappings = {
            "open": "visit status open",
            "temporarily_closed": "visit status temporary closed",
            "permanently_closed": "visit status permanently closed",
            "moved": "visit status moved",
            "pos_not_found": "visit status not found"
        }
        status_sums = {}
        for key, header in status_mappings.items():
            col = next((c for c in df.columns if str(c).strip().lower().replace('\n', ' ') == header), None)
            status_sums[key] = pd.to_numeric(df[col], errors='coerce').sum() if col else 0

        # 3. Date extraction
        actual_date_col = next((c for c in df.columns if str(c).strip().lower().replace('\n', ' ') == 'to date'), None)
        extracted_period = None
        if actual_date_col:
            dates = pd.to_datetime(df[actual_date_col], errors='coerce').dropna()
            if not dates.empty:
                extracted_period = dates.max().strftime('%d-%B-%Y')
            else:
                last_val = str(df[actual_date_col].dropna().iloc[-1]).strip()
                import re
                match = re.search(r'(\d{4})-(\d{2})-(\d{2})', last_val)
                if match:
                    extracted_period = pd.to_datetime(last_val).strftime('%d-%B-%Y')
                else:
                    extracted_period = last_val
        
        return total_sum, status_sums, extracted_period, None
    except Exception as e:
        return 0, {}, None, f"Error processing file: {str(e)}"

# Helper function to create final summary table
def create_final_summary(shinsa_counts, total_shinsa, mtd_total, mtd_status_sums):
    # Mapping table label to internal normalized status keys
    mapping = [
        ("Visit Status Open", "open"),
        ("Visit Status temporary closed", "temporarily_closed"),
        ("Visit Status permanently closed", "permanently_closed"),
        ("Visit status Moved", "moved"),
        ("Visit status Not found", "pos_not_found")
    ]
    
    data = []
    for label, status_key in mapping:
        # Shinsa counts
        count = 0
        if shinsa_counts is not None and not shinsa_counts.empty:
            match = shinsa_counts[shinsa_counts['Status'] == status_key]
            if not match.empty:
                count = match.iloc[0]['Count']
        
        # MTD counts for this category
        mtd_val = mtd_status_sums.get(status_key, 0)
        
        # Calculations based on standard audit logic
        visit_ach = (mtd_val / mtd_total) if mtd_total > 0 else 0.0
        audit_ach = (count / total_shinsa) if total_shinsa > 0 else 0.0
        row_coverage = (count / mtd_val) if mtd_val > 0 else 0.0
        
        data.append({
            "Visit Status": label,
            "Successful Visits": int(mtd_val),
            "Visit Ach%": f"{visit_ach:.2%}",
            "Audited Visits": int(count),
            "Audit Ach%": f"{audit_ach:.2%}",
            "Coverage %": f"{row_coverage:.2%}"
        })
    
    # Calculation for Grand Total row
    grand_total_coverage = (total_shinsa / mtd_total) if mtd_total > 0 else 0.0
    
    grand_total = {
        "Visit Status": "Grand Total",
        "Successful Visits": int(mtd_total),
        "Visit Ach%": "100.00%", 
        "Audited Visits": int(total_shinsa),
        "Audit Ach%": "100.00%",
        "Coverage %": f"{grand_total_coverage:.2%}"
    }
    
    return pd.DataFrame(data), grand_total

# Helper function to extract totals from DB dataframe
def extract_mtd_totals_from_df(df):
    try:
        # Total Successful Visits
        total_sum = pd.to_numeric(df['MTD Successful Visits'], errors='coerce').sum()
        
        # Individual Status Sums
        status_mappings = {
            "open": "Visit Status Open",
            "temporarily_closed": "Visit Status Temporary Closed",
            "permanently_closed": "Visit Status Permanently Closed",
            "moved": "Visit Status Moved",
            "pos_not_found": "Visit Status Not Found"
        }
        status_sums = {}
        for key, col_name in status_mappings.items():
            status_sums[key] = pd.to_numeric(df[col_name], errors='coerce').sum()
        
        # Period from 'To Date'
        extracted_period = None
        if 'To Date' in df.columns:
            extracted_period = pd.to_datetime(df['To Date']).max().strftime('%d-%B-%Y')
            
        return total_sum, status_sums, extracted_period
    except Exception as e:
        return 0, {}, None

# Initialize session state for DB-generated MTD results
if 'db_mtd_results' not in st.session_state:
    st.session_state.db_mtd_results = None

# Header
st.title("Audit Analyzer")
st.write("Upload reports or generate from DB to start the analysis.")

st.divider()

# Upload sections
col1, col2 = st.columns(2)

with col1:
    st.subheader("Audit report from Shinsa")
    shinsa_file = st.file_uploader(
        "Choose Shinsa file", 
        type=["xlsx", "xls", "csv"],
        key="shinsa_uploader",
        label_visibility="collapsed"
    )

with col2:
    st.subheader("MTD Successful visit Report")
    
    # Automation Section
    auto_col1, auto_col2 = st.columns(2)
    with auto_col1:
        f_date = st.date_input("From Date", value=datetime(2026, 1, 25))
    with auto_col2:
        t_date = st.date_input("To Date", value=datetime.now())
    
    if st.button("🚀 Generate MTD Report from DB", use_container_width=True):
        with st.spinner("Connecting to DB and generating report..."):
            f_date_str = f_date.strftime('%Y-%m-%d')
            t_date_str = t_date.strftime('%Y-%m-%d')
            t_date_filename = t_date.strftime('%d-%m-%Y')
            
            df_me_raw = fetch_db_data(get_me_mtd_query(f_date_str, t_date_str))
            df_sup_raw = fetch_db_data(get_sup_mtd_query(t_date_str))
            
            if not df_me_raw.empty:
                df_summary_auto = process_me_summary(df_me_raw)
                excel_bytes = generate_automation_excel(df_me_raw, df_summary_auto, df_sup_raw, t_date_filename)
                
                # Store aggregated totals in session state for the final table
                m_total, m_sums, m_period = extract_mtd_totals_from_df(df_me_raw)
                st.session_state.db_mtd_results = {
                    'total_sum': m_total,
                    'status_sums': m_sums,
                    'period': m_period,
                    'excel_bytes': excel_bytes,
                    'filename': f"National_ME_MTD_Report_{t_date_filename}.xlsx"
                }
                st.success("MTD Data loaded successfully from Database.")
            else:
                st.error("No data found for the selected date.")

    st.markdown("<p style='text-align: center; color: gray;'>— OR —</p>", unsafe_allow_html=True)
            
    mtd_file = st.file_uploader(
        "Upload existing MTD file", 
        type=["xlsx", "xls", "csv"],
        key="mtd_uploader",
        label_visibility="collapsed"
    )
    
    # Show script download if data was generated from DB
    if st.session_state.db_mtd_results:
        st.download_button(
            label="💾 Download Generated MTD Report",
            data=st.session_state.db_mtd_results['excel_bytes'],
            file_name=st.session_state.db_mtd_results['filename'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.divider()

# Variables to store analysis results
shinsa_status_df = None
shinsa_total = 0
mtd_total_sum = 0
mtd_status_sums = {}
auto_period = None
error = None

# Processing Shinsa (Always from upload)
if shinsa_file:
    with st.spinner("Analyzing Shinsa..."):
        shinsa_status_df, shinsa_total, error = get_shinsa_summary(shinsa_file)
        if error:
            st.error(error)

# Processing MTD (Priority: 1. Manual Upload, 2. Database Generation)
if mtd_file:
    with st.spinner("Analyzing MTD File..."):
        mtd_total_sum, mtd_status_sums, auto_period, error = get_mtd_summary(mtd_file)
        if error:
            st.error(error)
elif st.session_state.db_mtd_results:
    # Use silently loaded DB data
    mtd_total_sum = st.session_state.db_mtd_results['total_sum']
    mtd_status_sums = st.session_state.db_mtd_results['status_sums']
    auto_period = st.session_state.db_mtd_results['period']

# Helper function to export to Excel
import io
def export_to_excel(df, title):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Summary', startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['Summary']
        
        # Add title
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=14)
        
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        cell = worksheet.cell(row=1, column=1)
        cell.value = title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
        # Style header row of table
        table_header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        for col_num, value in enumerate(df.columns, 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.fill = table_header_fill
            cell.font = Font(bold=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Style data rows
        for row_num in range(3, len(df) + 3):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if row_num == len(df) + 2: # Grand Total row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for i, col in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 5)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return output.getvalue()

# Final Summary Table (Shows if we have Shinsa and ANY source of MTD)
if shinsa_status_df is not None and (mtd_file or st.session_state.db_mtd_results) and not error:
    # Use the extracted date from the 'To Date' column, or a placeholder if missing
    display_date = auto_period if auto_period else "Date Not Found"
    report_title = f"Audit Summary Report [{display_date}]"
    
    st.markdown(f"""
        <div style='background-color: #333333; padding: 10px; border-radius: 5px; margin-bottom: 20px; text-align: center;'>
            <h2 style='color: white; margin: 0;'>{report_title}</h2>
        </div>
    """, unsafe_allow_html=True)
    
    final_df, grand_total_row = create_final_summary(shinsa_status_df, shinsa_total, mtd_total_sum, mtd_status_sums)
    
    # Add grand total row to dataframe for display
    display_df = pd.concat([final_df, pd.DataFrame([grand_total_row])], ignore_index=True)
    
    # Custom styling for the summary table
    st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True
    )
    
    # Display as clean, professional text instead of a box
    st.markdown(f"### Final Coverage for {display_date}: **{grand_total_row['Coverage %']}**")

    # Download button directly below, aligned to the left
    col_dl, col_empty = st.columns([1.2, 2.8])
    with col_dl:
        excel_data = export_to_excel(display_df, report_title)
        st.download_button(
            label="📊 Download Excel Report",
            data=excel_data,
            file_name=f"Audit_Summary_{display_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# Overall status help
if not shinsa_file or (not mtd_file and not st.session_state.db_mtd_results):
    st.info("Please upload Shinsa report or generate MTD from DB to see the Summary Report.")
