import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter, column_index_from_string
from sqlalchemy import create_engine, text

# Database Parameters
DB_PARAMS = {
    "dbname": "gp_dev",
    "user": "report_user",
    "password": "report#Gp*User!#__D",
    "host": "gp-stg.cf44ysgum7u8.ap-southeast-1.rds.amazonaws.com",
    "port": 5432
}

# SQLAlchemy engine
DATABASE_URL = f"postgresql://{DB_PARAMS['user']}:{DB_PARAMS['password']}@{DB_PARAMS['host']}:{DB_PARAMS['port']}/{DB_PARAMS['dbname']}"
engine = create_engine(DATABASE_URL)

# Date Configuration
from_date = '2026-01-25'
custom_date = '2026-02-24'
custom_date_filename = datetime.strptime(custom_date, '%Y-%m-%d').strftime('%d-%m-%Y')

def me_mtd_rpt(custom_date):
    return text(f"""
    WITH report_data AS (
	SELECT
	    from_date,
	    to_date,
	    circle,
	    region,
	    cluster_name,
	    territory,
	    house,
	    dh_code,
	    me_code,
	    CASE WHEN successful_visit > 0 THEN 'Working' ELSE 'Absent' END AS status,
	   	mtd_working_day,
	    COALESCE(NULLIF(monthly_target, 0), 0) AS monthly_target,
	    COALESCE(NULLIF(open, 0), 0) AS open,
	    COALESCE(NULLIF(temporarily_closed, 0), 0) AS temporarily_closed,
	    COALESCE(NULLIF(permanently_closed, 0), 0) AS permanently_closed,
	    COALESCE(NULLIF(moved, 0), 0) AS moved,
	    COALESCE(NULLIF(not_found, 0), 0) AS not_found,
		avg_time_spent_in_between_pos_service,
	    ---Visit Data--
	    COALESCE(NULLIF(mtd_target, 0), 0) AS mtd_target,
	    COALESCE(NULLIF(mtd_served_days, 0), 0) AS mtd_served_days,
	    COALESCE(NULLIF(late_checkin, 0), 0) AS late_checkin,
	    COALESCE(NULLIF(successful_visit, 0), 0) AS successful_visit,
	    COALESCE(ROUND((successful_visit::numeric / NULLIF(mtd_target, 0)) * 100, 0), 0)::int || '%' AS successful_visit_ach,
		avg_time_spent_in_pos,
	    COALESCE(NULLIF(three_or_more_min_visit_count, 0), 0) AS three_or_more_min_visit_count,
	    '80%' AS grater_or_equ_three_min_visit_target_percent,
	    COALESCE(ROUND((three_or_more_min_visit_count ::numeric / NULLIF(mtd_target, 0)) * 100, 0), 0)::int || '%' AS less_equ_three_min_visit_ach,
	    --POS Wise Data--
	    COALESCE(NULLIF(monthly_unique_pos_visit_target, 0), 0) AS monthly_unique_pos_visit_target,
	    COALESCE(NULLIF(one_successful_visit_count_pos, 0), 0) AS one_successful_visit_count_pos,
	    COALESCE(NULLIF(two_successful_visit_count_pos, 0), 0) AS two_successful_visit_count_pos,
	    COALESCE(NULLIF(three_or_more_successful_visit_count_pos, 0),0) AS three_or_more_successful_visit_count_pos,
	    COALESCE(NULLIF(COALESCE(one_successful_visit_count_pos, 0) + COALESCE(two_successful_visit_count_pos, 0) + COALESCE(three_or_more_successful_visit_count_pos, 0), 0), 0) AS mtd_unique_pos_visit,
	    --Poster POS Data--
	    COALESCE(NULLIF(COALESCE(five_or_more_poster_pos, 0) + COALESCE(four_poster_pos, 0) + COALESCE(three_poster_pos, 0), 0), 0) AS three_or_more_poster_pos_count,
	    '90%' AS three_or_more_poster_pos_target,
	    COALESCE(NULLIF(five_or_more_poster_pos, 0), 0) AS five_or_more_poster_pos,
	    COALESCE(NULLIF(four_poster_pos, 0), 0) AS four_poster_pos,
	    COALESCE(NULLIF(three_poster_pos, 0), 0) AS three_poster_pos,
	    COALESCE(NULLIF(two_poster_pos, 0), 0) AS two_poster_pos,
	    COALESCE(NULLIF(one_poster_pos, 0), 0) AS one_poster_pos,
	    COALESCE(NULLIF(COALESCE(five_or_more_poster_pos, 0) + COALESCE(four_poster_pos, 0) + COALESCE(three_poster_pos, 0) + COALESCE(two_poster_pos, 0) + COALESCE(one_poster_pos, 0), 0), 0) AS all_poster_pos,
	    --Festoon POS Data--
	    COALESCE(NULLIF(COALESCE(three_or_more_festoon_pos, 0) + COALESCE(two_festoon_pos, 0) + COALESCE(one_festoon_pos, 0), 0), 0) AS atleast_one_festoon_pos_count,
	    '70%' AS festoon_pos_count_target,
	    COALESCE(NULLIF(three_or_more_festoon_pos, 0), 0) AS three_or_more_festoon_pos,
	    COALESCE(NULLIF(two_festoon_pos, 0), 0) AS two_festoon_pos,
	    COALESCE(NULLIF(one_festoon_pos, 0), 0) AS one_festoon_pos,
	    --Shopscreen pos--
	    COALESCE(NULLIF(shopscreen_pos, 0), 0) AS shopscreen_pos,
	    '30%' AS shopscreen_pos_count_target,
	    COALESCE(NULLIF(cover_sticker_pos, 0), 0) AS cover_sticker_pos,
	    '30%' AS cover_sticker_pos_count_target,
	    COALESCE(NULLIF(no_bag, 0), 0) AS no_bag,
	    COALESCE(NULLIF(damaged_bag, 0), 0) AS damaged_bag,
	    COALESCE(NULLIF(no_hammer, 0), 0) AS no_hammer,
	    COALESCE(NULLIF(no_pliers, 0), 0) AS no_pliers,
	    COALESCE(NULLIF(anti_cutter, 0), 0) AS anti_cutter,
	    COALESCE(NULLIF(duster, 0), 0) AS duster,
	    COALESCE(NULLIF(scotch_tape, 0), 0) AS scotch_tape,
	    COALESCE(NULLIF(glue, 0), 0) AS glue,
	    COALESCE(NULLIF(tar_or_string, 0), 0) AS tar_or_string,
	    COALESCE(NULLIF(board_pin, 0), 0) AS board_pin,
	    COALESCE(NULLIF(pin_or_perek, 0), 0) AS pin_or_perek
	from
	    dashboard_reporting.me_mtd_data
	WHERE
	    to_date = '{custom_date}'   
)
SELECT
	from_date AS "From Date",
    to_date AS "To Date",
    circle AS "Circle",
    region AS "Region",
    cluster_name AS "Cluster",
    territory AS "Territory",
    house AS "Distribution House",
    dh_code AS "DH Code",
    me_code AS "ME Code",
    status as "Status",
    mtd_working_day as "MTD Working Days Count",
    mtd_served_days AS "MTD Served Days",
    CASE 
    WHEN COALESCE(mtd_working_day, 0) - COALESCE(mtd_served_days, 0) < 0 THEN 0
    ELSE COALESCE(mtd_working_day, 0) - COALESCE(mtd_served_days, 0)
	END AS "Vacant Days",
    late_checkin AS "Late Check In Days Count [After 9 AM]",
    monthly_target AS "Monthly Visit Target",
    mtd_target AS "MTD Visit Target",
    successful_visit AS "MTD Successful Visits",
    successful_visit_ach as "Successful Visit Ach%",
    open AS "Visit Status Open",
    temporarily_closed AS "Visit Status Temporary Closed",
    permanently_closed AS "Visit Status Permanently Closed",
    moved AS "Visit Status Moved",
    not_found AS "Visit Status Not Found",
    avg_time_spent_in_pos AS "Avg Spent Time in POS",
    avg_time_spent_in_between_pos_service AS "Avg Time Spent in Between POS Service",
    three_or_more_min_visit_count AS ">= 3 Min Visit Count",
    grater_or_equ_three_min_visit_target_percent as ">= 3 Min Visit Target %",
    less_equ_three_min_visit_ach as ">= 3 Min Visit Ach%",
    --end--
    monthly_unique_pos_visit_target as  "Monthly Unique POS Visit Target",
    one_successful_visit_count_pos as "1 Successful Visit Count POS",
    two_successful_visit_count_pos as "2 Successful Visit Count POS",
    three_or_more_successful_visit_count_pos as "3&+ Successful Visit Count POS",
    mtd_unique_pos_visit as "MTD Unique POS Visited",
    COALESCE(ROUND((mtd_unique_pos_visit ::numeric / NULLIF(monthly_unique_pos_visit_target, 0)) * 100, 0), 0)::int || '%' AS "MTD Unique POS Visit Ach.%",
    -- Poster POS Data--
    three_or_more_poster_pos_count as ">= 3 Poster POS Count",
    three_or_more_poster_pos_target as ">= 3 Poster POS Count Target",
    COALESCE(ROUND((three_or_more_poster_pos_count ::numeric / NULLIF(monthly_unique_pos_visit_target, 0)) * 100, 0), 0)::int || '%' as ">= 3 Poster POS Count Ach%",
    five_or_more_poster_pos as "5&+ Poster POS",	
    four_poster_pos as "4 Poster POS",
    three_poster_pos as "3 Poster POS",	
    two_poster_pos as "2 Poster POS",
    one_poster_pos as "1 Poster POS",
    COALESCE(NULLIF(mtd_unique_pos_visit - all_poster_pos, 0), 0) AS "0 Poster POS",
    ---
    atleast_one_festoon_pos_count as "Atleast 1 Festoon POS Count",
    festoon_pos_count_target as "Festoon POS Count Target",
    COALESCE(ROUND((atleast_one_festoon_pos_count ::numeric / NULLIF(monthly_unique_pos_visit_target, 0)) * 100, 0), 0)::int || '%' "Festoon POS Count Ach%",
    three_or_more_festoon_pos as "3&+ Festoon POS",
    two_festoon_pos as "2 Festoon POS",
    one_festoon_pos as "1 Festoon POS",
    COALESCE(NULLIF(mtd_unique_pos_visit - atleast_one_festoon_pos_count, 0), 0) as "0 Festoon POS",
    shopscreen_pos as "Atleast 1 Shopscreen POS Count",
    shopscreen_pos_count_target as "Shopscreen POS Count Target",
    COALESCE(ROUND((shopscreen_pos ::numeric / NULLIF(monthly_unique_pos_visit_target, 0)) * 100, 0), 0)::int || '%' as "Shopscreen POS Count Ach%",
    cover_sticker_pos as "Atleast 1 Cover Sticker POS Count",
    cover_sticker_pos_count_target as "Cover Sticker POS Count Target",
    COALESCE(ROUND((cover_sticker_pos ::numeric / NULLIF(monthly_unique_pos_visit_target, 0)) * 100, 0), 0)::int || '%' as "Cover Sticker POS Count Ach%",
    no_bag AS "No Bag",
	damaged_bag AS "Damaged Bag",
	no_hammer AS "No Hammer",
	no_pliers AS "No Pliers",
	anti_cutter AS "Anti Cutter",
	duster AS "Duster",
	scotch_tape AS "Scotch Tape",
	glue AS "Glue",
	tar_or_string AS "Tar/String",
	board_pin AS "Board Pin",
	pin_or_perek AS "Pin/Perek"
FROM report_data
order by me_code;
    """)

def sup_mtd_rpt(custom_date):
    return text(f"""
    WITH report_data AS (
	SELECT
	    sm.from_date,
	    sm.to_date,
	    sm.circle,
	    sm.region,
	    sm.sup_code,
	    COALESCE(sm.mtd_working_day, 0) AS mtd_working_day,
	    COALESCE(sm.sup_served_days, 0) AS sup_served_days,
	    GREATEST(COALESCE(sm.mtd_working_day, 0) - COALESCE(sm.sup_served_days, 0), 0) AS vacant_days,
	    COALESCE(sm.late_check_in_count, 0) AS late_check_in_count,
	    sm.late_check_in_average_time,
	    COALESCE(sm.mtd_me_tag_target, 0) AS mtd_me_tag_target,
	    COALESCE(sm.me_tag_ach, 0) AS me_tag_ach,
	    COALESCE(ROUND((sm.me_tag_ach::numeric / NULLIF(sm.mtd_me_tag_target, 0)) * 100, 0), 0)::int || '%' AS joint_ach,
	    COALESCE(sm.mtd_working_day, 0) * 20 AS mtd_joint_visit_target,
	    COALESCE(sm.mtd_joint_visit_ach, 0) AS mtd_joint_visit_ach,
	    COALESCE(sm.successful_joint_visit_ach, 0) AS successful_joint_visit_ach,
	    COALESCE(sm.mtd_working_day, 0) * 10 AS mtd_solo_visit_terget,
	    COALESCE(sm.mtd_solo_visit_ach, 0) AS mtd_solo_visit_ach,
	    COALESCE(sm.visit_status_open, 0) AS visit_status_open,
	    COALESCE(sm.visit_status_same_pos_different_name, 0) AS visit_status_same_pos_different_name,
	    COALESCE(sm.mtd_unique_pos_visit, 0) AS mtd_unique_pos_visit,
	    COALESCE(sm.mtd_joint_visit_ach, 0) + COALESCE(sm.mtd_solo_visit_ach, 0) AS mtd_total_pos_visit,
	    COALESCE(sm.mtd_me_successful_visit, 0) AS mtd_me_successful_visit,
	    COALESCE(sm.mtd_me_unique_pos_visited, 0) AS mtd_me_unique_pos_visited,
	    COALESCE(sm.three_min_iseqgreater_visit_count, 0) AS three_min_iseqgreater_visit_count,
	    '80%' AS three_min_visit_target,
	    COALESCE(sm.three_iseqgreater_poster_pos_count, 0) AS three_iseqgreater_poster_pos_count,
	    '90%' AS poster_pos_count_target,
	    COALESCE(sm.five_plus_poster_pos, 0) AS five_plus_poster_pos,
	    COALESCE(sm.four_poster_pos, 0) AS four_poster_pos,
	    COALESCE(sm.three_poster_pos, 0) AS three_poster_pos,
	    COALESCE(sm.two_poster_pos, 0) AS two_poster_pos,
	    COALESCE(sm.one_poster_pos, 0) AS one_poster_pos,
	    COALESCE(sm.five_plus_poster_pos, 0) + 
	    COALESCE(sm.four_poster_pos, 0) + 
	    COALESCE(sm.three_poster_pos, 0) + 
	    COALESCE(sm.two_poster_pos, 0) + 
	    COALESCE(sm.one_poster_pos, 0) AS all_poster_pos,
	    '70%' AS festoon_pos_count_target,
	    COALESCE(sm.three_plus_festoon_pos, 0) AS three_plus_festoon_pos,
	    COALESCE(sm.two_festoon_pos, 0) AS two_festoon_pos,
	    COALESCE(sm.one_festoon_pos, 0) AS one_festoon_pos,
	    COALESCE(sm.three_plus_festoon_pos, 0) + 
	    COALESCE(sm.two_festoon_pos, 0) + 
	    COALESCE(sm.one_festoon_pos, 0) AS all_festoon_pos,
	    COALESCE(sm.atleast_one_shopscreen_pos, 0) AS atleast_one_shopscreen_pos,
	    '30%' AS shopscreen_pos_count_target,
	    COALESCE(sm.atleast_one_cover_sticker_pos, 0) AS atleast_one_cover_sticker_pos,
	    '30%' AS cover_sticker_pos_count_target
	FROM dashboard_reporting.sup_mtd_data sm
	WHERE sm.to_date = '{custom_date}'
)
SELECT
	from_date AS "Start Date",
	to_date AS "Till Date",
	circle AS "Circle",
	region AS "Region",
	sup_code AS "Sup Code",
	mtd_working_day AS "MTD Working Day Count",
	sup_served_days AS "SUP Served Days",
	vacant_days AS "Vacant Days",
	late_check_in_count AS "Late Check In Count",
	late_check_in_average_time AS "Late Check in Average Time",
	mtd_me_tag_target AS "Monthly ME Tag Target",
	me_tag_ach AS "ME Tag Ach",
	joint_ach AS "Ach %",
	mtd_joint_visit_target AS "MTD Joint Visit Target",
	mtd_joint_visit_ach AS "MTD Joint Visit Ach",
	successful_joint_visit_ach AS "MTD Successful Joint Visit Ach",
	COALESCE(ROUND((mtd_joint_visit_ach::numeric / NULLIF(mtd_joint_visit_target, 0)) * 100, 0), 0)::int || '%' AS "Final Ach %",
	mtd_solo_visit_terget AS "MTD Solo Visit Target",
	mtd_solo_visit_ach AS "MTD Solo Visit Ach",
	COALESCE(ROUND((mtd_solo_visit_ach::numeric / NULLIF(mtd_solo_visit_terget, 0)) * 100, 0), 0)::int || '%' AS "Ach %",
	visit_status_open AS "Visit Status Open",
	visit_status_same_pos_different_name AS "Visit Status Same POS Different Name",
	mtd_total_pos_visit AS "MTD Total POS Visit",
	mtd_unique_pos_visit AS "MTD Unique POS Visit",
	COALESCE(ROUND((mtd_unique_pos_visit::numeric / NULLIF(mtd_total_pos_visit, 0)) * 100, 0), 0)::int || '%' AS "MTD Unique POS Visit %",
	three_min_iseqgreater_visit_count AS ">= 3 Min Visit Count",
	three_min_visit_target AS ">= 3 Min Visit Target %",
	COALESCE(ROUND((three_min_iseqgreater_visit_count::numeric / NULLIF(mtd_me_successful_visit, 0)) * 100, 0), 0)::int || '%' AS ">= 3 Min Visit Ach%",
	three_iseqgreater_poster_pos_count AS ">= 3 Poster POS Count",
	poster_pos_count_target AS ">= 3 Poster POS Count Target",
	COALESCE(ROUND((three_iseqgreater_poster_pos_count::numeric / NULLIF(mtd_me_unique_pos_visited, 0)) * 100, 0), 0)::int || '%' AS ">= 3 Poster POS Count Ach%",
	five_plus_poster_pos AS "5&+ Poster POS",
	four_poster_pos AS "4 Poster POS",
	three_poster_pos AS "3 Poster POS",
	two_poster_pos AS "2 Poster POS",
	one_poster_pos AS "1 Poster POS",
	COALESCE(mtd_me_unique_pos_visited, 0) - COALESCE(all_poster_pos, 0) AS "0 Poster POS",
	all_festoon_pos AS "Atleast 1 Festoon POS Count",
	festoon_pos_count_target AS "Festoon POS Count Target",
	COALESCE(ROUND((all_festoon_pos::numeric / NULLIF(mtd_me_unique_pos_visited, 0)) * 100, 0), 0)::int || '%' AS "Festoon POS Count Ach%",
	three_plus_festoon_pos AS "3&+ Festoon POS",
	two_festoon_pos AS "2 Festoon POS",
	one_festoon_pos AS "1 Festoon POS",
	COALESCE(mtd_me_unique_pos_visited, 0) - COALESCE(all_festoon_pos, 0) AS "0 Festoon POS",
	atleast_one_shopscreen_pos AS "Atleast 1 Shopscreen POS Count",
	shopscreen_pos_count_target AS "Shopscreen POS Count Target",
	COALESCE(ROUND((atleast_one_shopscreen_pos::numeric / NULLIF(mtd_me_unique_pos_visited, 0)) * 100, 0), 0)::int || '%' AS "Shopscreen POS Count Ach%",
	atleast_one_cover_sticker_pos AS "Atleast 1 Cover Sticker POS Count",
	cover_sticker_pos_count_target AS "Cover Sticker POS Count Target",
	COALESCE(ROUND((atleast_one_cover_sticker_pos::numeric / NULLIF(mtd_me_unique_pos_visited, 0)) * 100, 0), 0)::int || '%' AS "Cover Sticker POS Count Ach%"
FROM report_data;
    """)

def fetch_raw_data(custom_date, query_function):
    try:
        query = query_function(custom_date)
        df = pd.read_sql_query(query, con=engine)
        return df
    except Exception as e:
        print(f"❌ Error fetching data: {e}")
        return pd.DataFrame()

def me_summary_data(df):
    if df.empty:
        return pd.DataFrame()

    circle_col = 'Circle'
    region_col = 'Region'
    assigned_me_col = 'ME Code'
    status_col = 'Status'

    summary_df = df[[circle_col, region_col]].drop_duplicates().sort_values(by=circle_col).reset_index(drop=True)

    summary_df['Assigned ME'] = summary_df[region_col].apply(
        lambda region: df[df[region_col] == region][assigned_me_col].nunique()
    )
    summary_df['Working ME'] = summary_df[region_col].apply(
        lambda region: df[(df[region_col] == region) & (df[status_col] == 'Working')][assigned_me_col].nunique()
    )

    status_fields_mapping = {
        'Visit Status Open': 'Open',
        'Visit Status Temporary Closed': 'Temporary closed',
        'Visit Status Permanently Closed': 'Permanently closed',
        'Visit Status Moved': 'Moved',
        'Visit Status Not Found': 'Not found'
    }

    for df_col, display_col in status_fields_mapping.items():
        summary_df[display_col] = summary_df[region_col].apply(
            lambda region: df[df[region_col] == region][df_col].sum()
        )

    summary_df['POS Visited'] = summary_df[list(status_fields_mapping.values())].sum(axis=1)
    summary_df['Open POS %'] = summary_df['Open'] / summary_df['POS Visited']

    unique_pos_col = 'MTD Unique POS Visited'
    if unique_pos_col in df.columns:
        summary_df['Unique POS Visited'] = summary_df[region_col].apply(
            lambda region: df[df[region_col] == region][unique_pos_col].sum()
        )
    else:
        summary_df['Unique POS Visited'] = 0

    poster_mapping = {
        'Min 1 Festoon': 'Atleast 1 Festoon POS Count',
        'Min 1 Shopscreen': 'Atleast 1 Shopscreen POS Count',
        'Cover Sticker': 'Atleast 1 Cover Sticker POS Count',
        '0 Poster': '0 Poster POS',
        '1 Poster': '1 Poster POS',
        '2 Poster': '2 Poster POS',
        '3 and + Poster': '>= 3 Poster POS Count'
    }

    for display_col, raw_col in poster_mapping.items():
        summary_df[display_col] = summary_df[region_col].apply(
            lambda region: df[df[region_col] == region][raw_col].sum() if raw_col in df.columns else 0
        )

    for col in poster_mapping.keys():
        percent_col = col + ' POS %'
        summary_df[percent_col] = summary_df[col] / summary_df['Unique POS Visited']

    final_columns = [
        'Circle', 'Region', 'Assigned ME', 'Working ME', 'POS Visited', 'Open',
        'Temporary closed', 'Permanently closed', 'Moved', 'Not found', 'Open POS %',
        'Unique POS Visited',
        'Min 1 Festoon', 'Min 1 Shopscreen', 'Cover Sticker',
        '0 Poster', '1 Poster', '2 Poster', '3 and + Poster',
        'Min 1 Festoon POS %', 'Min 1 Shopscreen POS %', 'Cover Sticker POS %',
        '0 Poster POS %', '1 Poster POS %', '2 Poster POS %', '3 and + Poster POS %'
    ]

    summary_df = summary_df[final_columns].copy()
    
    numeric_columns = [col for col in final_columns if col not in ['Circle', 'Region'] and not col.endswith('%')]
    percent_columns = [col for col in summary_df.columns if col.endswith('%')]

    total_row = {col: '' for col in final_columns}
    for col in numeric_columns:
        total_row[col] = summary_df[col].sum()

    for col in percent_columns:
        total_row[col] = f"{round(summary_df[col].mean() * 100):.0f} %"
        summary_df[col] = (summary_df[col] * 100).round(0).astype(int).astype(str) + " %"

    summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)
    return summary_df

def save_the_data(df_raw, df_summary, date_filename, filename_prefix="National ME MTD Report", df_sup=None):
    save_path = r"D:\GP Audit Team"
    folder_path = os.path.join(save_path, f"ME MTD Rpt {date_filename}")
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, f"{filename_prefix} {date_filename}.xlsx")

    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_raw.to_excel(writer, sheet_name="ME MTD RPT", index=False)
            df_summary.to_excel(writer, sheet_name="National Summary", index=False)
            if df_sup is not None and not df_sup.empty:
                df_sup.to_excel(writer, sheet_name="SUP MTD RPT", index=False)

        wb = load_workbook(file_path)
        thin = Side(border_style="thin", color="000000")
        all_border = Border(top=thin, bottom=thin, left=thin, right=thin)
        
        for sheet in wb.worksheets:
            sheet.row_dimensions[1].height = 60
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.border = all_border

            for cell in sheet[1]:
                col_letter = cell.column_letter 
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                cell.alignment = Alignment(wrap_text=True)
                sheet.column_dimensions[col_letter].width = 12

            if sheet.title == "ME MTD RPT":
                for col_idx in [18, 27, 28, 34, 36, 45, 37, 46, 52, 53, 55, 56]: # Percent cols
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if isinstance(cell.value, str) and '%' in cell.value:
                                try:
                                    cell.value = float(cell.value.replace('%', '')) / 100
                                    cell.number_format = '0%'
                                except: pass

            elif sheet.title == "National Summary":
                last_row = sheet.max_row
                for col_idx in [11] + list(range(20, 27)):
                    for row in sheet.iter_rows(min_row=2, max_row=last_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if isinstance(cell.value, str) and '%' in cell.value:
                                try:
                                    cell.value = float(cell.value.replace('%', '')) / 100
                                    cell.number_format = '0%'
                                except: pass
                
                # Apply conditional formatting
                for col_letter in ['K', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
                    col_range = f"{col_letter}2:{col_letter}{last_row-1}"
                    sheet.conditional_formatting.add(col_range, ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84', end_type='max', end_color='63BE7B'))

        wb.save(file_path)
        return file_path
    except Exception as e:
        print(f"❌ Error saving Excel: {e}")
        return None

if __name__ == "__main__":
    print(f"🚀 Starting National report generation for: {custom_date}")
    df_me = fetch_raw_data(custom_date, me_mtd_rpt)
    df_sup = fetch_raw_data(custom_date, sup_mtd_rpt)

    if not df_me.empty:
        df_summary = me_summary_data(df_me)
        path = save_the_data(df_me, df_summary, custom_date_filename, df_sup=df_sup)
        if path: print(f"✅ Report saved: {path}")
    else:
        print("❌ No data found.")